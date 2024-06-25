from openpyxl import load_workbook

def search_in_excel(file_path, search_text):
    wb = load_workbook(filename=file_path, read_only=True)
    results = []
    for sheet in wb:
        for row in sheet.iter_rows(min_row=2):
            if row[1].value and search_text.lower() in row[1].value.lower():
                results.append({'sheet': sheet.title, 'row': row[0].row})
    return results
def read_row_values(file_path, sheet_name, row_number):
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb[sheet_name]
    row_values = []
    for cell in sheet[row_number]:
        row_values.append(cell.value)
    return row_values